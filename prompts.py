CODE_GEN_PROMPT = """
Ты — Senior Python Developer. Напиши парсер для SPIL Matrix (Excel).

ВХОДНЫЕ ДАННЫЕ:
JSON-сэмпл (первые 60-100 строк).

ЗАДАЧА:
Напиши функцию def parse_excel(file_path: str, target_sheet: str = None) -> list[dict]:

ВАЖНОЕ ТРЕБОВАНИЕ БЕЗОПАСНОСТИ:
При вычислении индексов (например, col - 2) всегда оборачивай их в max(1, ...), чтобы избежать ошибки "Row or column values must be at least 1". Пример: range(max(1, WALL_COL - 2), ...).

АЛГОРИТМ (СТРОГО СЛЕДУЙ ЕМУ, НЕ МЕНЯЙ ЛОГИКУ):

1. ОТКРЫТИЕ ФАЙЛА:
   import openpyxl
   import re 
   wb = openpyxl.load_workbook(file_path, data_only=True)
   sheet = None

   if target_sheet and target_sheet in wb.sheetnames:
       sheet = wb[target_sheet]
   else:
       for s in wb.sheetnames:
           if "SPIL" in s.upper():
               sheet = wb[s]
               break
       if sheet is None:
           sheet = wb.active

2. ДИНАМИЧЕСКИЙ ПОИСК ЗАГОЛОВКОВ (SCANNER):
   - HEADER_ROW = None
   - Ключевые слова: ["DESCRIPTION", "MANUFACTURER", "PART NUMBER", "TOTAL", "QTY", "QUANTITY", "MEASURE", "U.O.M"]
   - Просканируй строки с 1 по 100.
   - Если >= 3 совпадений ключевых слов -> HEADER_ROW = row; break.
   - Если HEADER_ROW is None -> return []

   - WALL_COL = None
   - В HEADER_ROW ищем первую колонку с "DESCRIPTION", "MEASURE", "ITEM" или "POS" -> это WALL_COL.

3. ПОИСК СТРОК МЕТАДАННЫХ (TAGS, SERIAL, MODEL):
   - TAG_ROW_IDX = None; SERIAL_ROW_IDX = None; MODEL_ROW_IDX = None
   - Если HEADER_ROW is not None:
       - Сканируем строки row от 1 до HEADER_ROW (включительно).
       - row_str = строка в верхнем регистре.
       - Если TAG_ROW_IDX is None:
           if "EQUIPMENT" in row_str and "TAG" in row_str: TAG_ROW_IDX = row
       - Если "SERIAL" in row_str and "NUMBER" in row_str: SERIAL_ROW_IDX = row
       - Если "MODEL" in row_str: MODEL_ROW_IDX = row
       - Если TAG_ROW_IDX is None and "TAG" in row_str: TAG_ROW_IDX = row

   if HEADER_ROW is None or WALL_COL is None or TAG_ROW_IDX is None or SERIAL_ROW_IDX is None or MODEL_ROW_IDX is None:
       return []

4. ДИНАМИЧЕСКОЕ СОПОСТАВЛЕНИЕ КОЛОНОК (SMART MAPPING):
   - cols = {{}}
   - header_texts = {{}} 
   - start_scan_col = max(1, WALL_COL - 2)

   - Пройди по колонкам col от start_scan_col до sheet.max_column.
   - header_text = объединение строк (max(1, HEADER_ROW - 2)) до (HEADER_ROW + 3).
   - val = header_text.upper().replace("\\n", " ").replace("\\r", " ").strip()
   - while "  " in val: val = val.replace("  ", " ")
   - header_texts[col] = val

   - ЛОГИКА ЗАПОЛНЕНИЯ:
     if "MANUFACTURER NAME" in val: cols['manuf'] = col
     if "UNIT" in val and "PRICE" in val and "TOTAL" not in val: cols['price'] = col
     if "TRUE" in val and "MANUFACTURER" in val: cols['true_part'] = col
     if "MEASURE" in val or "U.O.M" in val: cols['uom'] = col

     if "DESCRIPTION" in val: cols['desc'] = col
     if "LEAD" in val: cols['lead'] = col
     if "WAREHOUSE" in val or "SAP PART" in val: cols['warehouse'] = col
     if ("CHECK" in val or "MATERIAL" in val) and "DESCRIPTION" not in val: cols['check'] = col
     if "CURRENCY" in val: cols['currency'] = col

     if 'manuf' not in cols and "MANUFACTURER" in val:
          forbidden = ["TRUE", "PART", "TOTAL", "RECOMMENDED", "QTY", "NO", "DWG", "ID"]
          if not any(bad_word in val for bad_word in forbidden):
              cols['manuf'] = col

5. ПОИСК И РАЗДЕЛЕНИЕ ТЕГОВ (FINAL ROBUST PARSER):
   - tag_map = {{}} 
   - last_valid_tags = None 

   # ---------------------------------------------------------------------
   # ФУНКЦИЯ parse_tag_cell (УМНАЯ СКЛЕЙКА И РАЗДЕЛЕНИЕ)
   # ---------------------------------------------------------------------
   # def parse_tag_cell(cell_value) -> list[str]:
   #    if not cell_value: return []
   #    val_str = str(cell_value).strip()
   #    val_str = val_str.replace(" -", "-").replace("- ", "-")
   #
   #    # ВАЖНО: Лечим разрыв слэшей из-за переносов строк (22A/B\n/C -> 22A/B/C)
   #    val_str = re.sub(r"\\s*/\\s*", "/", val_str)
   #    
   #    # 1. Заменяем переносы, запятые и ДВА И БОЛЕЕ пробела на пайпы (|)
   #    val_str = re.sub(r"[\\n\\r\\t,;&]", "|", val_str)
   #    val_str = re.sub(r"\\s{{2,}}", "|", val_str)
   #    
   #    # 2. Склеиваем суффиксы (A, B, 01), если перед ними 1 пробел, а после - пайп, СЛЭШ или конец строки
   #    val_str = re.sub(r"(?<=\\w) (?=\\w{{1,2}}(?:[|/]|$))", "", val_str)
   #    
   #    # 3. Оставшиеся одиночные пробелы тоже считаем разделителями
   #    val_str = val_str.replace(" ", "|")
   #    
   #    raw_tokens = [t.strip() for t in val_str.split("|") if t.strip()]
   #    expanded_tags = []
   #    
   #    def is_strong(part):
   #         return (len(part) >= 4 and any(c.isdigit() for c in part)) or \
   #                (len(part) >= 3 and part[0].isalpha() and any(c.isdigit() for c in part))
   #
   #    for token in raw_tokens:
   #        token = token.lstrip("-")
   #        parts = token.split('-')
   #        for i in range(len(parts) - 1):
   #            if is_strong(parts[i]) and is_strong(parts[i+1]):
   #                parts[i] += '|'
   #        token = "-".join(parts).replace("|-", "|")
   #        
   #        for sub in token.split('|'):
   #            if '/' not in sub: expanded_tags.append(sub)
   #            else:
   #                s_parts = [x.strip() for x in sub.split('/') if x.strip()]
   #                if not s_parts: continue
   #                base = s_parts[0]
   #                expanded_tags.append(base)
   #                for suffix in s_parts[1:]:
   #                    if '-' in suffix: expanded_tags.append(suffix)
   #                    elif suffix.isdigit(): base = re.sub(r"\\d+$", suffix, base); expanded_tags.append(base)
   #                    elif len(suffix) == 1 and base[-1].isalpha(): base = base[:-1] + suffix; expanded_tags.append(base)
   #                    else: expanded_tags.append(suffix)
   #    return expanded_tags
   # ---------------------------------------------------------------------

   - Иди по колонкам col от 1 до WALL_COL (строго < WALL_COL).
   - val_raw = sheet.cell(TAG_ROW_IDX, col).value
   - val_str = str(val_raw).strip() if val_raw else ""

   - skip_col = False
   - if any(x in val_str.upper() for x in ["DATE", "REV", "VENDOR", "SHEET", "ORDER", "PROJECT"]):
         last_valid_tags = None; skip_col = True

   - Если not skip_col:
       if not val_str and last_valid_tags:
           tag_map[col] = last_valid_tags
       elif val_str:
           tags = parse_tag_cell(val_str) 
           valid_tags = [t for t in tags if not (len(t) < 3 and t.isdigit())]
           if valid_tags:
               tag_map[col] = valid_tags
               last_valid_tags = valid_tags
           else:
               last_valid_tags = None

6. СБОР ДАННЫХ (SMART MERGE С ПОДДЕРЖКОЙ MULTI-SERIAL):
   - DATA_START = HEADER_ROW + 1
   - Если "desc" не найден в cols -> return []
   - last_created_items = [] 

   - lead_is_weeks = False
   - if 'lead' in cols:
         ht = header_texts.get(cols['lead'], "").upper()
         if "WEEK" in ht: lead_is_weeks = True

   # ---------------------------------------------------------------------
   # ФУНКЦИЯ РАЗДЕЛЕНИЯ МЕТАДАННЫХ (SERIAL / MODEL)
   # ---------------------------------------------------------------------
   # def split_metadata_cell(cell_val, count_needed):
   #     if not cell_val: return [None] * count_needed
   #     s_val = str(cell_val).strip()
   #     parts = [p.strip() for p in re.split(r'[\\n\\r]+', s_val) if p.strip()]
   #     if len(parts) == count_needed: return parts
   #     parts = [p.strip() for p in re.split(r'\\s{{2,}}', s_val) if p.strip()]
   #     if len(parts) == count_needed: return parts
   #     return [s_val] * count_needed
   # ---------------------------------------------------------------------

   - Иди от DATA_START до sheet.max_row.
   - desc_val = sheet.cell(row, cols["desc"]).value
   - desc_str = str(desc_val).strip() if desc_val else ""

   - if not desc_str: continue
   - if "TOTAL" in desc_str.upper() or "PAGE" in desc_str.upper(): last_created_items = []; continue
   - clean_desc = desc_str.replace(".0", "").replace(".", "")
   - if clean_desc.isdigit(): continue

   - row_has_qty = False
   - current_row_items = []

   - has_metadata = False
   - if 'manuf' in cols and sheet.cell(row, cols['manuf']).value: has_metadata = True
   - if 'uom' in cols and sheet.cell(row, cols['uom']).value: has_metadata = True

   - Иди по элементам tag_map (col, tags_list):
     - qty_val = sheet.cell(row, col).value
     - try: qty = float(qty_val) except: qty = 0

     - Если qty > 0:
       - row_has_qty = True
       - serial_raw = sheet.cell(SERIAL_ROW_IDX, col).value if SERIAL_ROW_IDX else None
       # serials_expanded = split_metadata_cell(serial_raw, len(tags_list))
       - model = sheet.cell(MODEL_ROW_IDX, col).value if MODEL_ROW_IDX else None

       curr = "USD"
       if 'currency' in cols: curr = sheet.cell(row, cols['currency']).value
       elif 'price' in cols:
           price_header = header_texts.get(cols['price'], "")
           if "EUR" in price_header: curr = "EUR"
           elif "USD" in price_header: curr = "USD"

       lead_val = sheet.cell(row, cols['lead']).value if 'lead' in cols else None
       if lead_val is not None and lead_is_weeks:
           try: lead_val = float(lead_val) * 7
           except: pass 

       for i, tag_name in enumerate(tags_list):
           item = {{
               "equipment_tag": tag_name, 
               "serial_number": serials_expanded[i],   
               "model": model,            
               "quantity": qty,          
               "description": desc_str, 
               "uom": sheet.cell(row, cols['uom']).value if 'uom' in cols else None,
               "manufacturer": sheet.cell(row, cols['manuf']).value if 'manuf' in cols else None,
               "true_part_number": sheet.cell(row, cols['true_part']).value if 'true_part' in cols else None,
               "unit_price": sheet.cell(row, cols['price']).value if 'price' in cols else None,
               "warehouse_number": sheet.cell(row, cols['warehouse']).value if 'warehouse' in cols else None,
               "lead_time": lead_val,
               "material_check": sheet.cell(row, cols['check']).value if 'check' in cols else None,
               "currency": curr
           }}
           data.append(item)
           current_row_items.append(item)

   - Если row_has_qty == True:
       last_created_items = current_row_items

   - Иначе (Если row_has_qty == False):
       if last_created_items and desc_str and not has_metadata:
           for item in last_created_items:
               item["description"] += " " + desc_str
       elif has_metadata:
           last_created_items = []

7. ВОЗВРАТ: List[Dict]

ВХОДНОЙ JSON:
{context_data}
"""