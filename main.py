import os
import json
import time
import pandas as pd
import openpyxl
import importlib.util
from llm_service import CodeGeneratorService

# --- НАСТРОЙКИ ---
INPUT_FILE = "J04-2122-TM-PK-7500115894-001 PNEUMATIC CONVEYING PLANT (COPERION).xlsx"
OUTPUT_FOLDER = "gpt_parser"
SCRIPTS_FOLDER = "gpt_parsers"  # <-- НОВАЯ ПАПКА ДЛЯ СКРИПТОВ ГПТ


def get_excel_snippet(file_path, sheet_name=None, rows_count=60):
    """Читает первые N строк конкретного листа"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    snippet_data = {}
    row_count = 0
    for row in ws.iter_rows(min_row=1, max_row=rows_count):
        row_has_data = False
        for cell in row:
            if cell.value is not None:
                val = str(cell.value).strip()
                if val:
                    key = f"Row {cell.row}, Col {cell.column_letter}"
                    snippet_data[key] = val
                    row_has_data = True
        if row_has_data:
            row_count += 1

    return json.dumps(snippet_data, ensure_ascii=False), row_count


def main():
    start_time = time.time()
    print(f"🚀 ЗАПУСК: Обработка файла {INPUT_FILE}...")

    if not os.path.exists(INPUT_FILE):
        print(f"❌ Файл {INPUT_FILE} не найден")
        return

    if not os.path.exists(OUTPUT_FOLDER):
        os.makedirs(OUTPUT_FOLDER)

    # Создаем папку для скриптов, если ее нет
    if not os.path.exists(SCRIPTS_FOLDER):
        os.makedirs(SCRIPTS_FOLDER)
        print(f"📁 Создана папка для сгенерированных кодов: {SCRIPTS_FOLDER}")

    base_name = os.path.splitext(os.path.basename(INPUT_FILE))[0]
    final_output_path = os.path.join(OUTPUT_FOLDER, f"{base_name}-gpt_parser.xlsx")

    # 1. ОТКРЫВАЕМ ФАЙЛ, ЧТОБЫ НАЙТИ ЛИСТЫ
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=False)
    visible_sheets = []
    for sheet in wb:
        if sheet.sheet_state == 'visible':
            visible_sheets.append(sheet.title)
        else:
            print(f"👻 Скрытый лист пропущен: [{sheet.title}]")
    wb.close()

    print(f"📄 Очередь листов: {visible_sheets}")

    all_data_merged = []
    service = CodeGeneratorService()

    # 2. БЕЖИМ ПО ЛИСТАМ
    for sheet_name in visible_sheets:
        print(f"\n--- Лист: [{sheet_name}] ---")

        snippet_json, rows_with_data = get_excel_snippet(INPUT_FILE, sheet_name=sheet_name)

        if rows_with_data < 5:
            print("⏩ Пропуск (мало данных)")
            continue

        print(f"📦 Генерирую парсер...")

        parser_obj = service.generate_parser_code(snippet_json)

        if not parser_obj.python_code:
            print("❌ GPT не вернул код.")
            continue

        # Сохраняем скрипт В НОВУЮ ПАПКУ
        safe_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '_', '-')).strip()
        script_filename = f"parser_{safe_name}.py"
        script_path = os.path.join(SCRIPTS_FOLDER, script_filename)  # <-- ПУТЬ СОХРАНЕНИЯ

        code_content = parser_obj.python_code.replace("```python", "").replace("```", "")
        with open(script_path, "w", encoding="utf-8") as f:
            f.write(code_content)

        # Запускаем скрипт ИЗ НОВОЙ ПАПКИ
        try:
            spec = importlib.util.spec_from_file_location(f"mod_{safe_name}", script_path)
            module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(module)

            if hasattr(module, "parse_excel"):
                sheet_data = module.parse_excel(INPUT_FILE, target_sheet=sheet_name)

                if sheet_data:
                    print(f"✅ УСПЕХ: Найдено {len(sheet_data)} записей.")
                    all_data_merged.extend(sheet_data)
                else:
                    print(f"⚠️ Пустой результат (скрипт не нашел структуру SPIL).")
            else:
                print("❌ Нет функции parse_excel.")

        except Exception as e:
            print(f"❌ Ошибка выполнения: {e}")

    # 3. ФИНАЛ
    print("\n" + "=" * 30)
    if all_data_merged:
        print(f"📊 ИТОГО: {len(all_data_merged)} записей.")

        df = pd.DataFrame(all_data_merged)

        rename_map = {
            "equipment_tag": "Equipment Tag Number",
            "serial_number": "Serial Number",
            "model": "Model",
            "quantity": "Quantity",
            "uom": "UNIT OF MEASURE",
            "description": "DESCRIPTION OF PART",
            "manufacturer": "MANUFACTURER NAME",
            "true_part_number": "TRUE MANUFACTURER'S PART NUMBER",
            "material_check": "MATERIAL CHECK ASTM JIS DIN",
            "lead_time": 'LEADTIME READY TO SHIP AFTER RECEIPT OF ORDER "ARO" (days)',
            "unit_price": "UNIT PRICE",
            "currency": "CURRENCY",
            "warehouse_number": "Company Warehouse Number. (Company SAP Part No.) To Be Provided by Company"
        }
        df = df.rename(columns=rename_map)

        desired_order = [
            "Equipment Tag Number",
            "Serial Number",
            "Model",
            "Quantity",
            "UNIT OF MEASURE",
            "DESCRIPTION OF PART",
            "MANUFACTURER NAME",
            "TRUE MANUFACTURER'S PART NUMBER",
            "MATERIAL CHECK ASTM JIS DIN",
            'LEADTIME READY TO SHIP AFTER RECEIPT OF ORDER "ARO" (days)',
            "UNIT PRICE",
            "CURRENCY",
            "Company Warehouse Number. (Company SAP Part No.) To Be Provided by Company"
        ]

        for col in desired_order:
            if col not in df.columns: df[col] = ""
        df = df[desired_order]

        df.to_excel(final_output_path, index=False)
        print(f"🎉 Результат сохранен: {final_output_path}")
    else:
        print("⚠️ Нет данных для сохранения.")

    print(f"⏱️ Время: {time.time() - start_time:.2f} сек")


if __name__ == "__main__":
    main()